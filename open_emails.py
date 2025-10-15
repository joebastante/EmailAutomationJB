import openpyxl
import platform
import subprocess
import urllib.parse
from openai import OpenAI
import os

def split_name(full_name):
    """Split full name into first and last name."""
    parts = full_name.strip().split()
    if len(parts) == 0:
        return "", ""
    elif len(parts) == 1:
        return parts[0], ""
    else:
        return parts[0], " ".join(parts[1:])

def generate_email_body(client, first_name, email_type, context):
    """Generate email body using OpenAI API."""
    prompt = f"""I'd like you to craft an email to check in a professional connection and colleague. The email should be about 3 to 4 sentences in total to greet them, ask how the are doing, and check in on them. I want them to know I hope they are well I'm here to support them if ever needed. {context}. The email should be started with "Hi {first_name}." on its own line followed by a blank line and then the email text The tone should be casual but professional. Wording should be natural and sincere. This person has a relationship value of "{email_type}". Use the relationship types below to adjust the tone based on the depth of relationship I have with each person.
The relationship types are as follows:
-          1: This email is for those with whom I'm friends and have known for quite a while.
-          2: This email is for those with whom I've worked and have a personal relationship.
-          3: This email is for people I know professionally with whom I've had occasional interactions.
-          4: This email is for those I've spoken to only a couple times and I'm trying to build a connection."""
    
    try:
        response = client.chat.completions.create(
            model="gpt-5-mini",
            messages=[
                {"role": "user", "content": prompt}
            ]
            #temperature=0.7
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"Error generating email: {e}")
        return None

def open_outlook_email(recipient_email, subject, body):
    """Open new Outlook email window (cross-platform)."""
    # Encode the body for URL
    encoded_body = urllib.parse.quote(body)
    encoded_subject = urllib.parse.quote(subject)
    
    # Create mailto link
    mailto_link = f"mailto:{recipient_email}?subject={encoded_subject}&body={encoded_body}"
    
    system = platform.system()
    
    try:
        if system == "Windows":
            # Try to open with Outlook specifically on Windows
            try:
                subprocess.run(["outlook", f"/c", "ipm.note", f"/m", recipient_email], 
                             check=False)
            except:
                # Fallback to default mail client
                os.startfile(mailto_link)
        elif system == "Darwin":  # macOS
            subprocess.run(["open", mailto_link], check=True)
        else:  # Linux and other Unix-like systems
            subprocess.run(["xdg-open", mailto_link], check=True)
        
        return True
    except Exception as e:
        print(f"Error opening email client: {e}")
        return False

def process_excel_workbook(workbook_path, api_key):
    """Process Excel workbook and generate emails."""
    # Initialize OpenAI client
    client = OpenAI(api_key=api_key)
    
    # Load workbook
    try:
        wb = openpyxl.load_workbook(workbook_path)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return
    
    # Get "Working" worksheet
    if "Working" not in wb.sheetnames:
        print("Error: 'Working' worksheet not found in workbook")
        return
    
    ws = wb["Working"]
    
    # Process each row starting from row 2
    row_num = 2
    while True:
        # Check if row has data (check column A)
        full_name_cell = ws[f"A{row_num}"].value
        if full_name_cell is None or str(full_name_cell).strip() == "":
            break  # No more data
        
        # Extract data from columns
        full_name = str(full_name_cell).strip()
        email_type = str(ws[f"B{row_num}"].value or "").strip()
        recipient_email = str(ws[f"C{row_num}"].value or "").strip()
        context = str(ws[f"D{row_num}"].value or "").strip()
        
        # Validate required fields
        if not recipient_email or not email_type:
            print(f"Row {row_num}: Missing required data (email or type), skipping...")
            row_num += 1
            continue
        
        # Split name
        first_name, last_name = split_name(full_name)
        
        print(f"\nProcessing row {row_num}: {full_name} ({recipient_email})")
        
        # Generate email body
        email_body = generate_email_body(client, first_name, email_type, context)
        
        if email_body is None:
            print(f"Row {row_num}: Failed to generate email, skipping...")
            row_num += 1
            continue
        
        # Open Outlook with new email
        subject = "Checking in"
        success = open_outlook_email(recipient_email, subject, email_body)
        
        if success:
            print(f"Row {row_num}: Email window opened successfully")
        else:
            print(f"Row {row_num}: Failed to open email window")
        
        # Pause to allow user to review email before moving to next
        #input("Press Enter to continue to next email...")
        
        row_num += 1
    
    print(f"\nProcessing complete. Processed {row_num - 2} rows.")

def main():
    # Configuration
    workbook_path = r"C:\Users\josep\OneDrive\shared\files\dynamic\busdocs\me\Networking_Plan.xlsm"  # Change this to your workbook path
    
    # Get OpenAI API key from environment variable
    api_key = os.environ.get("OPENAI_API_KEY")
    
    if not api_key:
        print("Error: OPENAI_API_KEY environment variable not set")
        print("Please set it using: export OPENAI_API_KEY='your-api-key' (Mac/Linux)")
        print("or: set OPENAI_API_KEY=your-api-key (Windows)")
        return
    
    process_excel_workbook(workbook_path, api_key)

if __name__ == "__main__":
    main()
